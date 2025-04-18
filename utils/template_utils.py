import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import tempfile
from datetime import datetime

def download_guru_template(subkriteria_list=None):
    """Membuat template Excel untuk import data guru"""
    try:
        wb = Workbook()
        
        # Sheet Data Guru
        ws_guru = wb.active
        ws_guru.title = "Data_Guru"
        ws_guru.append(['nama_guru', 'nip', 'jabatan', 'tanggal_masuk'])
        ws_guru.append(['Contoh Guru', '123456', 'Guru Matematika', '2023-01-01'])
        
        # Sheet Data Nilai (jika ada subkriteria)
        if subkriteria_list:
            ws_nilai = wb.create_sheet("Data_Nilai")
            headers = ['nip', 'tanggal_penilaian'] + [sub['nama_subkriteria'] for sub in subkriteria_list]
            ws_nilai.append(headers)
            ws_nilai.append(['123456', datetime.now().strftime('%Y-%m-%d')] + [3]*len(subkriteria_list))
        
        # Simpan ke file temporer
        import tempfile
        temp_path = tempfile.mktemp(suffix='.xlsx')
        wb.save(temp_path)
        
        return temp_path
        
    except Exception as e:
        raise Exception(f"Gagal membuat template: {str(e)}")

def download_nilai_template(guru_list, subkriteria_list):
    """
    Membuat template Excel untuk import nilai guru
    
    Parameters:
        guru_list: List data guru dari database
        subkriteria_list: List subkriteria dari database
    
    Returns:
        Path file template yang sudah dibuat
    """
    try:
        # Buat workbook baru
        wb = Workbook()
        
        # ===== Sheet Petunjuk =====
        ws_guide = wb.active
        ws_guide.title = "Petunjuk"
        
        # Judul petunjuk
        ws_guide.append(["PETUNJUK PENGISIAN TEMPLATE NILAI GURU"])
        ws_guide.merge_cells('A1:D1')
        ws_guide['A1'].font = Font(bold=True, size=14)
        ws_guide['A1'].alignment = Alignment(horizontal='center')
        
        # Isi petunjuk
        petunjuk = [
            ["1. Jangan mengubah struktur kolom yang sudah ada"],
            ["2. Gunakan NIP yang sudah terdaftar"],
            ["3. Nilai menggunakan skala 1-5 (1=Terendah, 5=Tertinggi)"],
            ["4. Format tanggal: YYYY-MM-DD"],
            ["5. Kolom kuning: Wajib diisi"],
            ["6. Kolom hijau: Otomatis terisi"],
            ["7. Hanya edit bagian nilai subkriteria"]
        ]
        
        for p in petunjuk:
            ws_guide.append(p)
        
        # ===== Sheet Data Nilai =====
        ws_nilai = wb.create_sheet("Data_Nilai")
        
        # Header
        headers = ["NIP", "Nama Guru", "Tanggal Penilaian"] + [sub['nama_subkriteria'] for sub in subkriteria_list]
        ws_nilai.append(headers)
        
        # Contoh data (3 baris)
        for i in range(min(3, len(guru_list))):
            guru = guru_list[i]
            row = [
                guru['nip'],
                guru['nama_guru'],
                datetime.now().strftime('%Y-%m-%d')
            ] + [""] * len(subkriteria_list)  # Kolom nilai dikosongkan
            ws_nilai.append(row)
        
        # ===== Styling =====
        # Style header
        header_style = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        
        for col in range(1, len(headers)+1):
            cell = ws_nilai.cell(row=1, column=col)
            cell.font = header_style
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        # Style kolom wajib (NIP dan Tanggal)
        wajib_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        for row in range(2, 5):
            ws_nilai.cell(row=row, column=1).fill = wajib_fill  # NIP
            ws_nilai.cell(row=row, column=3).fill = wajib_fill  # Tanggal
        
        # Style kolom nilai (warna berbeda)
        nilai_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        for col in range(4, len(headers)+1):
            for row in range(2, 5):
                ws_nilai.cell(row=row, column=col).fill = nilai_fill
        
        # Auto adjust column width
        for col in ws_nilai.columns:
            max_length = 0
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws_nilai.column_dimensions[col[0].column_letter].width = adjusted_width
        
        # Simpan ke file temporer
        temp_path = tempfile.mktemp(suffix='.xlsx')
        wb.save(temp_path)
        
        return temp_path
        
    except Exception as e:
        raise Exception(f"Gagal membuat template: {str(e)}")
    finally:
        if 'wb' in locals():
            wb.close()
     